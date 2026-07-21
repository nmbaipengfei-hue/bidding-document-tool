"""
Excel 文件解析模块
"""
import logging
from pathlib import Path
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class ExcelParser:
    """Excel 文件解析器"""
    
    def __init__(self, file_path):
        """
        初始化 Excel 解析器
        
        Args:
            file_path: Excel 文件路径
        """
        self.file_path = Path(file_path)
        self.workbook = None
        self.content = {}
        
    def parse(self):
        """
        解析 Excel 文件
        
        Returns:
            dict: 解析结果，包含所有工作表的数据
        """
        try:
            self.workbook = load_workbook(self.file_path)
            
            self.content = {}
            for sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]
                sheet_data = []
                
                for row in sheet.iter_rows(values_only=True):
                    sheet_data.append(list(row))
                
                self.content[sheet_name] = sheet_data
            
            logger.info(f"成功解析 Excel 文件: {self.file_path}")
            return {
                'success': True,
                'file': str(self.file_path),
                'sheets': list(self.workbook.sheetnames),
                'content': self.content
            }
        except Exception as e:
            logger.error(f"Excel 解析失败: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }
    
    def get_sheet_data(self, sheet_name):
        """
        获取指定工作表的数据
        
        Args:
            sheet_name: 工作表名称
            
        Returns:
            list: 表格数据
        """
        return self.content.get(sheet_name, [])
    
    def get_all_data(self):
        """
        获取所有工作表的数据
        
        Returns:
            dict: 所有工作表数据
        """
        return self.content
    
    def extract_keywords(self, keywords):
        """
        从 Excel 中提取关键词相关内容
        
        Args:
            keywords: 关键词列表
            
        Returns:
            dict: 关键词及其所在的位置和值
        """
        results = {}
        
        for sheet_name, sheet_data in self.content.items():
            for row_idx, row in enumerate(sheet_data):
                for col_idx, cell in enumerate(row):
                    if cell and isinstance(cell, str):
                        for keyword in keywords:
                            if keyword.lower() in str(cell).lower():
                                if keyword not in results:
                                    results[keyword] = []
                                results[keyword].append({
                                    'sheet': sheet_name,
                                    'row': row_idx,
                                    'col': col_idx,
                                    'value': str(cell)
                                })
        
        return results